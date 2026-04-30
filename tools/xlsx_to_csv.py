"""Convert all ASCEND Spring 2026 .xlsx workbooks (and .docx tables) to CSV.

Each worksheet -> one CSV file in matlab_ASCEND_S26/data/raw/.
Filename: <workbook>__<sheet>.csv  (spaces -> underscores).
"""
from __future__ import annotations
import csv, os, re, sys
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\P\Downloads\ASCENDSPRING26DATA")
DST = Path(__file__).resolve().parent.parent / "data" / "raw"
DST.mkdir(parents=True, exist_ok=True)


def slug(s: str) -> str:
    s = re.sub(r"\s+", "_", s.strip())
    return re.sub(r"[^A-Za-z0-9_.\-]", "", s)


def export_xlsx(path: Path) -> list[Path]:
    out: list[Path] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        name = slug(path.stem) + "__" + slug(ws.title) + ".csv"
        target = DST / name
        with target.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                w.writerow(["" if v is None else v for v in row])
        out.append(target)
        print(f"  -> {target.name}")
    return out


def main() -> int:
    if not SRC.exists():
        print(f"Source not found: {SRC}", file=sys.stderr)
        return 1
    for x in sorted(SRC.glob("*.xlsx")):
        print(f"[xlsx] {x.name}")
        export_xlsx(x)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
